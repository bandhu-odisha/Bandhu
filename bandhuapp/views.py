import os
import re
from datetime import datetime, timedelta

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from django.db.models import F
from django.http import HttpResponseRedirect, Http404, JsonResponse
from django.middleware.csrf import get_token
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_GET
from django.template import RequestContext
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import reverse
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode

from social_django.models import UserSocialAuth
from accounts.context_processors import pop_login_modal_flash
from accounts.models import User
from accounts.tokens import account_activation_token
from applications.anandakendra.models import Event as KendraEvent
from applications.ankurayan.models import Activity as AnkurayanActivity
from applications.ashram.models import Event as AshramEvent
from applications.charitywork.models import Activity as CharityActivity
from .models import (
    Designation, PeoplesDesignation, Profile, Photo, Initiatives, AboutUs,
    Mission, Staff, StaffExperience, StaffExperiencePhoto, Video, Volunteer,
    Gallery, Contact, HomePage, HomeVisitor, UrlData, CurrentUpdates, RecentActivity,
)
from .templatetags import permissions as temp_perms  # Template permissions
from .helpers import enrich_video_durations, people_card_from_assignments, proper_case
from .notice_links import resolve_notice_url

from openpyxl import Workbook
from openpyxl.styles import Font
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail


def _safe_login_next(next_value):
    """Allow only same-site relative paths for post-login ?next= and modal redirects."""
    if not next_value or not isinstance(next_value, str):
        return '/'
    n = next_value.strip()
    if not n.startswith('/') or n.startswith('//'):
        return '/'
    if n.startswith('/accounts/login'):
        return '/'
    return n


def _must_complete_member_profile(user):
    """
    Regular members need a Profile row before using the public home / landing API.
    Staff and superusers (is_staff / is_admin on this project's User model) are exempt
    so demo admins and operators are not forced through the signup-style profile flow.
    """
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_staff', False) or getattr(user, 'is_admin', False):
        return False
    return not Profile.objects.filter(user=user).exists()


def _can_manage_staff_experiences(user, staff=None):
    """Admins may manage experiences on any profile; the profile owner only on their own."""
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_admin', False) or getattr(user, 'is_staff', False):
        return True
    if staff is not None:
        profile = getattr(staff, 'profile', None)
        if profile is not None:
            return profile.user_id == user.pk
    return False


_PEOPLE_TAB_SLUGS = frozenset({'all', 'core-team', 'office-bearers'})


def _designation_to_people_tab_slug(title):
    return (title or '').strip().lower().replace(' ', '-')


def _people_back_url(request, staff_designations=None):
    """Link back to People page on the tab the visitor came from."""
    tab = _designation_to_people_tab_slug(request.GET.get('from', ''))
    if tab not in _PEOPLE_TAB_SLUGS:
        tab = 'all'
        if staff_designations:
            titles = {pd.designation.title for pd in staff_designations}
            if titles == {'Core Team'}:
                tab = 'core-team'
            elif titles and titles <= {'Office Bearers', 'Other'}:
                tab = 'office-bearers'
    return reverse('people_page') + f'#{tab}'


# Create your views here.

def index(request):
    if _must_complete_member_profile(request.user):
        messages.error(request, "Complete your Profile first.")
        return redirect('profile_page')

    # Visitor's Count
    if not request.session.get('home_page_visited', False):
        request.session['home_page_visited'] = True
        HomePage.objects.all().update(visitors_count=F('visitors_count') + 1)

    recent_events = []
    recent_events.extend(KendraEvent.objects.order_by('-date')[:10])
    recent_events.extend(AnkurayanActivity.objects.order_by('-date')[:10])
    recent_events.extend(AshramEvent.objects.order_by('-date')[:10])
    recent_events.extend(CharityActivity.objects.order_by('-date')[:10])

    recent_events.sort(key=lambda act: act.date, reverse=True)
    recent_events = recent_events[:10]

    context = {
        'initiatives': Initiatives.objects.all().first(),
        'about': AboutUs.objects.all().first(),
        'mission': Mission.objects.all().first(),
        'recent_events': recent_events,
        'volunteer': Volunteer.objects.all().first(),
        'photos': Photo.objects.filter(approved=True).order_by('-created'),
        'unapproved_photos': Photo.objects.filter(approved=False).order_by('created'),
        'curr_date': datetime.now().date(),
        'seven_day_delta': datetime.now().date() - timedelta(days=7),
        'content': HomePage.objects.all().first(),
        'current_updates': CurrentUpdates.objects.all()[:10],
        "people_designations": Designation.objects.all().values("title"),
        'videos': Video.objects.all().order_by('-created_at')[:10],
        'contact': Contact.objects.all().first(),
    }
    return render(request, 'landing_page.html', context)


def _build_landing_data(request):
    """Build JSON-serializable dict of landing page data for React frontend."""
    def url(path):
        if not path:
            return None
        if path.startswith(('http://', 'https://')):
            return path
        if path.startswith('/'):
            return path
        return request.build_absolute_uri(path) if request else f'/{path.lstrip("/")}'

    def file_url(field):
        if not field:
            return None
        try:
            if getattr(field, 'path', None) and os.path.isfile(field.path):
                return url(field.url)
        except (ValueError, AttributeError):
            pass
        name = getattr(field, 'name', None) or ''
        basename = os.path.basename(name)
        for img_dir in (
            os.path.join(settings.BASE_DIR, 'static', 'img'),
            os.path.join(settings.BASE_DIR, 'img'),
        ):
            if basename and os.path.isfile(os.path.join(img_dir, basename)):
                return url(static(f'img/{basename}'))
        return None

    def landing_logo_url():
        return url(static('img/bandhu-logo-navbar.png'))

    def resolve_media_or_static(relative_path, static_name=None):
        static_name = static_name or os.path.basename(relative_path)
        relative_path = relative_path.replace('\\', '/')
        media_path = os.path.join(settings.MEDIA_ROOT, *relative_path.split('/'))
        if os.path.isfile(media_path):
            media_url = settings.MEDIA_URL.lstrip('/')
            return url(f'/{media_url}/{relative_path}'.replace('//', '/'))
        for img_dir in (
            os.path.join(settings.BASE_DIR, 'static', 'img'),
            os.path.join(settings.BASE_DIR, 'img'),
        ):
            bundled = os.path.join(img_dir, static_name)
            if os.path.isfile(bundled):
                return url(static(f'img/{static_name}'))
        return None

    recent_events = []
    for e in KendraEvent.objects.order_by('-date')[:10]:
        recent_events.append({
            'id': e.id, 'name': e.name, 'description': getattr(e, 'description', '') or '',
            'date': e.date.isoformat(), 'thumb': file_url(e.thumb),
            'source': getattr(e.kendra, 'name', 'Anandakendra'),
        })
    for e in AnkurayanActivity.objects.order_by('-date')[:10]:
        recent_events.append({
            'id': f'a-{e.id}', 'name': e.name, 'description': getattr(e, 'description', '') or '',
            'date': e.date.isoformat(), 'thumb': file_url(getattr(e, 'thumb', None)),
            'source': 'Ankurayan',
        })
    for e in AshramEvent.objects.order_by('-date')[:10]:
        recent_events.append({
            'id': f'ash-{e.id}', 'name': e.name, 'description': getattr(e, 'description', '') or '',
            'date': e.date.isoformat(), 'thumb': file_url(getattr(e, 'thumb', None)),
            'source': 'Bandhughar',
        })
    for e in CharityActivity.objects.order_by('-date')[:10]:
        recent_events.append({
            'id': f'c-{e.id}', 'name': e.name, 'description': getattr(e, 'description', '') or '',
            'date': e.date.isoformat(), 'thumb': file_url(getattr(e, 'thumb', None)),
            'source': 'Other Activities',
        })
    recent_events.sort(key=lambda x: x['date'], reverse=True)
    recent_events = recent_events[:10]

    initiatives = Initiatives.objects.all().first()
    about = AboutUs.objects.all().first()
    mission = Mission.objects.all().first()
    content = HomePage.objects.all().first()
    volunteer = Volunteer.objects.all().first()
    gallery = Gallery.objects.all().first()
    contact = Contact.objects.all().first()

    def mission_carousel(qs):
        return [{'picture': file_url(p.picture)} for p in (qs or []) if file_url(p.picture)]

    data = {
        'initiatives': None,
        'about': None,
        'mission': None,
        'recent_events': recent_events,
        'volunteer': None,
        'photos': [],
        'content': None,
        'banner_image': None,
        'current_updates': [],
        'people_designations': list(Designation.objects.all().values_list('title', flat=True)),
        'videos': [],
        'contact': None,
        'recent_activities': [],
        'urls': {
            'login': reverse('login'),
            'signup': reverse('signup'),
            'ankurayan': reverse('ankurayan:ankurayan'),
            'anandakendra': reverse('anandakendra:anandakendra'),
            'ashram': reverse('ashram:ashram'),
            'charity_work': reverse('charitywork:charity_work'),
            'publications': reverse('publications:index'),
            'people': reverse('people_page'),
            'home': reverse('home'),
            'sanskar': reverse('pillar_sanskar'),
            'swaraj': reverse('pillar_swaraj'),
            'swabalamban': reverse('pillar_swabalamban'),
        },
        'user': {
            'is_authenticated': request.user.is_authenticated,
            'is_admin': getattr(request.user, 'is_admin', False),
        },
        'logo_url': landing_logo_url(),
        'about_slides': [],
        'profile_photos': [],
        'visitors': [],
    }

    for visitor in HomeVisitor.objects.filter(is_published=True):
        data['visitors'].append({
            'id': visitor.id,
            'name': visitor.name,
            'occupation': visitor.occupation,
            'place': visitor.place,
            'avatar': visitor.avatar,
            'about': visitor.about or '',
            'quote': visitor.quote,
            'facebookUrl': visitor.facebook_url or '',
            'linkedinUrl': visitor.linkedin_url or '',
            'photoUrl': file_url(visitor.photo),
        })

    if initiatives:
        data['initiatives'] = {
            'ankurayan_thumb': file_url(initiatives.ankurayan_thumb),
            'ankurayan_desc': initiatives.ankurayan_desc,
            'kendra_thumb': file_url(initiatives.kendra_thumb),
            'kendra_desc': initiatives.kendra_desc,
            'bandhughar_thumb': file_url(initiatives.bandhughar_thumb),
            'bandhughar_desc': initiatives.bandhughar_desc,
            'otheract_thumb': file_url(initiatives.otheract_thumb),
            'otheract_desc': initiatives.otheract_desc,
            'publications_thumb': file_url(initiatives.publications_thumb),
            'publications_desc': initiatives.publications_desc,
        }
    if about:
        data['about'] = {'tagline': about.tagline, 'desc': about.desc}
    if mission:
        data['mission'] = {
            'sanskar_tagline': mission.sanskar_tagline, 'sanskar_desc': mission.sanskar_desc,
            'sanskar_images': mission_carousel(mission.sanskarcarousel_set.all()),
            'swaraj_tagline': mission.swaraj_tagline, 'swaraj_desc': mission.swaraj_desc,
            'swaraj_images': mission_carousel(mission.swarajcarousel_set.all()),
            'swabalamban_tagline': mission.swabalamban_tagline, 'swabalamban_desc': mission.swabalamban_desc,
            'swabalamban_images': mission_carousel(mission.swabalambancarousel_set.all()),
        }
    if volunteer:
        data['volunteer'] = {'title': volunteer.title, 'tagline': volunteer.tagline}
    if content:
        banner_url = file_url(content.banner_image)
        data['content'] = {'banner_image': banner_url}
        data['banner_image'] = banner_url  # Hero uses this as first image on home
    if gallery:
        data['gallery_tagline'] = gallery.tagline

    about_slide_specs = (
        ('bandhuapp/gallery/about-slide-1-gardenia.png', 'about-slide-1-gardenia.png', 'Gardenia blossoms on campus.'),
        ('bandhuapp/gallery/about-slide-2-hibiscus.png', 'about-slide-2-hibiscus.png', 'Hibiscus in the campus gardens.'),
        ('bandhuapp/gallery/about-slide-3-campus.png', 'about-slide-3-campus.png', 'The Bandhu campus at dusk.'),
        ('bandhuapp/gallery/about-slide-blossoms.png', 'about-slide-blossoms.png', 'Ixora blooms along the walkways.'),
        ('bandhuapp/gallery/about-slide-4-ixora.png', 'about-slide-4-ixora.png', 'Seasonal blossoms across the grounds.'),
    )
    for relative_path, static_name, caption in about_slide_specs:
        slide_url = resolve_media_or_static(relative_path, static_name)
        if slide_url:
            data['about_slides'].append({'src': slide_url, 'caption': caption})

    hero_extra_specs = (
        ('bandhuapp/swaraj/our_mission1.jpg', 'our_mission1.jpg'),
        ('bandhuapp/gallery/about-slide-3-campus.png', 'about-slide-3-campus.png'),
        ('bandhuapp/gallery/about-slide-blossoms.png', 'about-slide-blossoms.png'),
    )
    reserved_gallery_names = {
        os.path.basename(relative_path).lower()
        for relative_path, _static_name, _caption in about_slide_specs
    }
    reserved_gallery_names.update(
        os.path.basename(static_name).lower()
        for _relative_path, static_name in hero_extra_specs
    )
    reserved_gallery_names.add('our_mission.jpg')

    data['hero_photos'] = []
    for relative_path, static_name in hero_extra_specs:
        hero_url = resolve_media_or_static(relative_path, static_name)
        if hero_url:
            data['hero_photos'].append({'picture': hero_url})

    seen_gallery_urls = set()
    seen_gallery_stems = set()
    for p in Photo.objects.filter(approved=True).order_by('-created'):
        picture_name = p.picture.name.replace('\\', '/')
        pic_data = {
            'picture': file_url(p.picture),
            'caption': p.caption or '',
            'tags': (p.tags or '').strip().split(),
        }
        if not pic_data['picture']:
            continue

        if picture_name.startswith('main_page/initiatives/'):
            continue

        if picture_name.startswith('profile_photos/'):
            data['profile_photos'].append(pic_data)
            continue

        if 'about-slide' in picture_name.lower():
            continue

        if not picture_name.startswith('bandhuapp/gallery/'):
            continue

        basename = os.path.basename(picture_name).lower()
        if basename in reserved_gallery_names:
            continue

        stem = re.sub(r'_[a-z0-9]{6,7}(?=\.)', '', os.path.splitext(basename)[0])
        if pic_data['picture'] in seen_gallery_urls or stem in seen_gallery_stems:
            continue
        seen_gallery_urls.add(pic_data['picture'])
        seen_gallery_stems.add(stem)
        data['photos'].append(pic_data)
    for u in CurrentUpdates.objects.all()[:10]:
        data['current_updates'].append({
            'desc': u.desc,
            'url': resolve_notice_url(u.desc, u.url) or None,
        })
    def youtube_video_id(script):
        if not script or not isinstance(script, str):
            return None
        m = re.search(r'(?:youtube\.com/embed/|youtube\.com/v/)([a-zA-Z0-9_-]{11})', script)
        if m:
            return m.group(1)
        m = re.search(r'youtube\.com/watch\?.*v=([a-zA-Z0-9_-]{11})', script)
        if m:
            return m.group(1)
        m = re.search(r'youtu\.be/([a-zA-Z0-9_-]{11})', script)
        if m:
            return m.group(1)
        m = re.search(r'[\?&]v=([a-zA-Z0-9_-]{11})', script)
        if m:
            return m.group(1)
        return None

    for v in Video.objects.all().order_by('-created_at')[:10]:
        vid = youtube_video_id(v.script)
        vd = getattr(v, 'duration', None)
        if vd is not None and not isinstance(vd, str):
            vd = str(vd)
        vd = (vd or '').strip() or None
        data['videos'].append({
            'title': v.title,
            'script': v.script,
            'video_id': vid,
            'duration': vd,
        })
    enrich_video_durations(data['videos'])
    if contact:
        data['contact'] = {
            'address': contact.address, 'contact_no': contact.contact_no, 'email': contact.email,
            'facebook_link': contact.facebook_link, 'twitter_link': contact.twitter_link,
        }
    for a in RecentActivity.objects.all().order_by('-start_date', '-date_created'):
        data['recent_activities'].append({
            'id': a.id, 'title': a.title, 'description': a.description,
            'link': resolve_notice_url(
                f'{a.title} {a.description}', a.link
            ) or '#',
            'start_date': a.start_date.isoformat() if a.start_date else None,
            'end_date': a.end_date.isoformat() if a.end_date else None,
        })
    return data


@require_GET
def landing_api(request):
    """JSON endpoint for React landing page data."""
    if _must_complete_member_profile(request.user):
        return JsonResponse({'error': 'Complete your profile first'}, status=400)
    data = _build_landing_data(request)
    return JsonResponse(data)


def index_react(request):
    """Serve the modern React landing page (same context as index, data passed as JSON)."""
    if _must_complete_member_profile(request.user):
        messages.error(request, "Complete your Profile first.")
        return redirect('profile_page')
    if not request.session.get('home_page_visited', False):
        request.session['home_page_visited'] = True
        HomePage.objects.all().update(visitors_count=F('visitors_count') + 1)
    flash = pop_login_modal_flash(request)
    request._login_modal_flash_consumed = True
    data = _build_landing_data(request)
    data['auth_modal'] = {
        'open_from_url': request.GET.get('login_modal') == '1',
        'err_code': flash['login_modal_err_code'],
        'prefill_email': flash['login_modal_prefill_email'],
        'next': _safe_login_next(request.GET.get('next')),
    }
    data['csrf_token'] = get_token(request)
    return render(
        request,
        'landing_react.html',
        {
            'landing_data': data,
            'login_modal_next': _safe_login_next(request.GET.get('next')),
        },
    )


def _pillar_context(mission, tagline_attr, desc_attr, images_queryset):
    """Build context for a pillar page (Sanskar, Swaraj, Swabalamban)."""
    content = HomePage.objects.all().first()
    if not mission:
        return {'title': '', 'tagline': '', 'desc': '', 'images': [], 'content': content}
    images = [p.picture for p in (images_queryset or [])]
    return {
        'title': getattr(mission, tagline_attr, '') or '',
        'tagline': getattr(mission, tagline_attr, '') or '',
        'desc': getattr(mission, desc_attr, '') or '',
        'images': images,
        'content': HomePage.objects.all().first(),
    }


def _sanskar_hero_picture(carousel_qs):
    """Prefer the Sanskar collage hero; fall back to the first carousel image."""
    for item in carousel_qs:
        name = (getattr(item.picture, 'name', None) or '').lower()
        if 'collage' in name:
            return item.picture
    return carousel_qs[0].picture if carousel_qs else None


def pillar_sanskar(request):
    mission = Mission.objects.all().first()
    carousel = list(mission.sanskarcarousel_set.all()) if mission else []
    ctx = _pillar_context(
        mission, 'sanskar_tagline', 'sanskar_desc',
        carousel)
    ctx['page_title'] = 'Sanskar'
    ctx['related_links'] = [
        {'name': 'Anandakendra', 'url': reverse('anandakendra:anandakendra')},
        {'name': 'Ankurayan', 'url': reverse('ankurayan:ankurayan')},
    ]
    ctx['sanskar_hero_image'] = _sanskar_hero_picture(carousel)
    ctx['pillar_hero_image'] = ctx['sanskar_hero_image']
    return render(request, 'pillar_page.html', ctx)


def pillar_swaraj(request):
    mission = Mission.objects.all().first()
    ctx = _pillar_context(
        mission, 'swaraj_tagline', 'swaraj_desc',
        mission.swarajcarousel_set.all() if mission else [])
    ctx['page_title'] = 'Swaraj'
    return render(request, 'pillar_page.html', ctx)


@login_required
def profile_page(request):
    user = request.user
    profile = Profile.objects.filter(user=user)

    if profile.exists():
        profile = profile[0]
        first_time = False
    else:
        # Dummy object to pass into the template
        profile = Profile(user=request.user)
        # If profile is being filled for the first time
        first_time = True

    if request.method == 'POST':
        # Required fields (all mandatory except street_address2)
        required = {
            'first_name': 'First name',
            'last_name': 'Last name',
            'gender': 'Gender',
            'dob': 'Date of birth',
            'profession': 'Profession',
            'contact_no': 'Contact number',
            'street_address1': 'Address',
            'city': 'City',
            'state': 'State',
            'pincode': 'PIN code',
        }
        errors = []
        for key, label in required.items():
            val = (request.POST.get(key) or '').strip()
            if not val:
                errors.append(f'{label} is required.')
        if first_time and not request.FILES.get('profile_pic'):
            errors.append('Profile picture is required.')
        if errors:
            messages.error(request, ' '.join(errors))
            # Re-populate profile from POST so user doesn't lose input
            profile.first_name = request.POST.get('first_name', '')
            profile.last_name = request.POST.get('last_name', '')
            profile.gender = request.POST.get('gender', 'M')
            profile.dob = request.POST.get('dob') or None
            profile.profession = request.POST.get('profession', '')
            profile.contact_no = request.POST.get('contact_no', '')
            profile.street_address1 = request.POST.get('street_address1', '')
            profile.street_address2 = request.POST.get('street_address2', '')
            profile.city = request.POST.get('city', '')
            profile.state = request.POST.get('state', '')
            profile.pincode = request.POST.get('pincode', '')
            context = {
                'profile': profile,
                'first_time': first_time,
                'content': HomePage.objects.all().first(),
            }
            return render(request, 'profile.html', context)

        profile.first_name = proper_case(request.POST['first_name'].strip())
        profile.last_name = proper_case(request.POST['last_name'].strip())
        profile.gender = request.POST['gender']
        profile.dob = request.POST['dob']
        profile.profession = proper_case(request.POST['profession'].strip())
        profile.contact_no = request.POST['contact_no'].strip()
        profile.street_address1 = proper_case(request.POST['street_address1'].strip())
        profile.street_address2 = proper_case((request.POST.get('street_address2') or '').strip())
        profile.city = proper_case(request.POST['city'].strip())
        profile.state = proper_case(request.POST['state'].strip())
        profile.pincode = request.POST['pincode'].strip()

        if 'profile_pic' in request.FILES:
            if profile.profile_pic and getattr(profile.profile_pic, 'name', None) and profile.profile_pic.name != 'profile_photos/man.png':
                profile.profile_pic.delete(False)
            profile.profile_pic = request.FILES['profile_pic']

        profile.save()

        if first_time:
            # Notify Admin for New User Sign Up
            current_site = get_current_site(request)

            from_email = settings.SENDER_EMAIL
            mail_subject = '[noreply] New User Signed Up'
            msg = 'A new User has signed up.'
            message = render_to_string('account_verification_email.html', {
                'user': user,
                'profile': profile,
                'domain': current_site.domain,
                'msg':msg,
                'uid':urlsafe_base64_encode(force_bytes(user.pk)),
                'token':account_activation_token.make_token(user),
            })
            to_email = settings.ADMINS_EMAIL
            # email = EmailMessage(
            #     mail_subject, message, from_email, to_email,
            # )
            # email.content_subtype = "html"
            # email.send()

            email = Mail(
                from_email=from_email,
                to_emails=to_email,
                subject=mail_subject,
                html_content=message,
            )
            try:
                sg = SendGridAPIClient(settings.SENDGRID_API_KEY)
                response = sg.send(email)
                print(response.status_code)
                print(response.body)
                print(response.headers)
            except Exception as e:
                print(e)

            logout(request)
            return redirect('account_activated')

        return HttpResponseRedirect('/profile/')

    context = {
        'profile': profile,
        'first_time': first_time,
        'content': HomePage.objects.all().first(),
    }
    return render(request,'profile.html', context)

@login_required
@user_passes_test(temp_perms.is_admin, redirect_field_name=None, login_url='/accounts/login/')
def add_image(request):
    """Only admin can add images to the site."""
    if request.method == 'POST':
        picture = request.FILES['image']
        caption = request.POST['caption']
        tag_list = request.POST.getlist('tags')

        tags = ""
        for tag in tag_list:
            tags += f'{tag} '

        Photo.objects.create(picture=picture, caption=caption, tags=tags, approved=True)
        messages.success(request, "Image added successfully!")

    return HttpResponseRedirect('/')

@login_required
def approve_image(request):
    if request.method == 'POST':
        image_pk = request.POST.get('image')
        status = request.POST.get('status')

        photo = get_object_or_404(Photo, pk=int(image_pk))

        if status == "approve":
            photo.approved = True
            photo.save()
            return JsonResponse({'response': 'approved'})
        else:
            photo.picture.delete(save=False)
            photo.delete()
            return JsonResponse({'response': 'discarded'})


    return HttpResponseRedirect('/')

@login_required
def extract_user_data(request):
    profiles = Profile.objects.order_by('-user__is_admin')

    file_path = settings.MEDIA_ROOT + '/sheets/user_profile_data.xlsx'
    excel = Workbook()
    sheet = excel.active

    # excel = openpyxl.load_workbook(filename = file_path)
    # sheet = excel.active
    col_width = [8, 30, 20, 8, 15, 15, 13, 30, 10, 20, 10, 10]

    sheet.cell(row=1, column=1).value = 'S.No'
    sheet.cell(row=1, column=2).value = 'Email'
    sheet.cell(row=1, column=3).value = 'Name'
    sheet.cell(row=1, column=4).value = 'Gender'
    sheet.cell(row=1, column=5).value = 'Date Of Birth'
    sheet.cell(row=1, column=6).value = 'Profession'
    sheet.cell(row=1, column=7).value = 'Contact'
    sheet.cell(row=1, column=8).value = 'Address'
    sheet.cell(row=1, column=9).value = 'City'
    sheet.cell(row=1, column=10).value = 'State'
    sheet.cell(row=1, column=11).value = 'Pincode'
    sheet.cell(row=1, column=12).value = 'Admin'

    for i in range(1, 13):
        sheet.cell(row = 1, column = i).font = Font(size=12, bold=True)
        sheet.column_dimensions[chr(65+(i-1))].width=col_width[i-1]

    for row, profile in enumerate(profiles, 2):
        sheet.cell(row=row, column=1).value = row - 1
        sheet.cell(row=row, column=2).value = profile.user.email
        sheet.cell(row=row, column=3).value = profile.get_full_name
        sheet.cell(row=row, column=4).value = profile.gender
        sheet.cell(row=row, column=5).value = profile.dob
        sheet.cell(row=row, column=6).value = profile.profession
        sheet.cell(row=row, column=7).value = profile.contact_no
        sheet.cell(row=row, column=8).value = f'{profile.street_address1}, {profile.street_address2}'
        sheet.cell(row=row, column=9).value = profile.city
        sheet.cell(row=row, column=10).value = profile.state
        sheet.cell(row=row, column=11).value = int(profile.pincode)
        sheet.cell(row=row, column=12).value = profile.user.is_admin

    excel.save(file_path)
    return HttpResponseRedirect(settings.MEDIA_URL + '/sheets/user_profile_data.xlsx')

def external_link(request,hash):
    if request.method == "GET":
        try:
            url_data = get_object_or_404(UrlData,pk=hash)
            url_data.times_followed += 1
            url_data.save()
            return HttpResponseRedirect(url_data.url)
        except Exception as e:
            return HttpResponseRedirect('/')

    return HttpResponseRedirect('/')

def people(request):
    if request.method == "GET":
        query_set = (
            PeoplesDesignation.objects.select_related(
                "staff", "staff__profile", "designation", "role"
            )
            .order_by("designation__rank", "rank")
            .all()
        )
        dict = {"All": []}
        all_by_staff = {}
        for i in query_set:
            if dict.get(i.designation.title, None) is None:
                dict[i.designation.title] = []
            dict[i.designation.title].append(people_card_from_assignments([i]))
            all_by_staff.setdefault(i.staff_id, []).append(i)
        dict["All"] = [
            people_card_from_assignments(assignments)
            for assignments in all_by_staff.values()
        ]
        # "Office Bearers" tab: use existing data or fall back to legacy "Other" designation
        if "Office Bearers" not in dict and "Other" in dict:
            dict["Office Bearers"] = dict.pop("Other")
        for key in list(dict.keys()):
            if key != "All" and not dict[key]:
                dict.pop(key)
        order = ["All", "Core Team", "Office Bearers"]
        rest = [k for k in dict if k not in order]
        data = {k: dict[k] for k in order if k in dict}
        for k in rest:
            data[k] = dict[k]
        return render(request, "people.html", {"data": data, "page_title": "People"})


def staff_profile(request, id):
    if request.method == "GET":
        staff_data = (
            Staff.objects.select_related("profile")
            .prefetch_related(
                "designations__designation",
                "designations__role",
                "experiences__photos",
            )
            .get(id=id)
        )
        experiences = staff_data.experiences.all()
        staff_designations = list(
            staff_data.designations.order_by("designation__rank", "rank")
        )
        designation_titles = [pd.designation.title for pd in staff_designations]
        primary_designation = staff_designations[0] if staff_designations else None
        office_bearer_assignments = [
            pd
            for pd in staff_designations
            if pd.designation.title in ("Office Bearers", "Other")
        ]
        is_office_bearer = bool(office_bearer_assignments)
        profession = (staff_data.profile.profession or "").strip()
        office_role_titles = [
            pd.role.title for pd in office_bearer_assignments if pd.role_id
        ]
        if office_role_titles:
            staff_position = " \u00b7 ".join(office_role_titles)
            primary_office = office_bearer_assignments[0]
            staff_occupation = (
                (primary_office.desc or "").strip() or profession
            )
        elif is_office_bearer and ", " in profession:
            parts = profession.rsplit(", ", 1)
            staff_position = parts[0].strip()
            staff_occupation = parts[1].strip() if len(parts) > 1 else ""
        else:
            staff_position = ""
            staff_occupation = profession
        designation_desc = " ".join(
            pd.desc.strip()
            for pd in staff_designations
            if pd.desc and pd.desc.strip()
        )
        return render(
            request,
            "staff-profile.html",
            {
                "staff": staff_data,
                "experiences": experiences,
                "designation_titles": designation_titles,
                "primary_designation": primary_designation,
                "designation_desc": designation_desc,
                "is_office_bearer": is_office_bearer,
                "staff_position": staff_position,
                "staff_occupation": staff_occupation,
                "can_manage_staff_experiences": _can_manage_staff_experiences(
                    request.user, staff_data
                ),
                "people_back_url": _people_back_url(
                    request, staff_designations=staff_designations
                ),
                "page_title": "Staff Profile",
            },
        )


def staff_experiences(request, id):
    """Dedicated page listing submitted experiences for a staff member (separate from profile)."""
    staff_data = get_object_or_404(
        Staff.objects.select_related("profile")
        .prefetch_related("designations__designation", "experiences__photos"),
        id=id,
    )
    experiences = staff_data.experiences.all()
    return render(
        request,
        "staff-experiences.html",
        {
            "staff": staff_data,
            "experiences": experiences,
        },
    )


def staff_share_experience(request, id):
    """Accept and process 'Share my experience' form from staff profile."""
    staff = get_object_or_404(Staff.objects.select_related('profile'), id=id)
    if not _can_manage_staff_experiences(request.user, staff):
        messages.error(request, 'You do not have permission to share experiences on this profile.')
        return redirect('staff_profile', id=id)
    if request.method == "POST":
        message = request.POST.get("experience", "").strip()
        if message:
            experience = StaffExperience.objects.create(staff=staff, message=message)
            captions = request.POST.getlist("photos_captions")
            for i, f in enumerate(request.FILES.getlist("photos")):
                if f.content_type and f.content_type.startswith("image/"):
                    raw_caption = captions[i].strip() if i < len(captions) else ""
                    caption = proper_case(raw_caption) if raw_caption else ""
                    StaffExperiencePhoto.objects.create(experience=experience, image=f, caption=caption)
            messages.success(request, "Thank you for sharing your experience!")
        else:
            messages.warning(request, "Please add your experience or thoughts before submitting.")
        return redirect(reverse("staff_profile", kwargs={"id": id}) + "#experiences-heading")
    return redirect("staff_profile", id=id)


def staff_edit_experience(request, id, experience_id):
    """Update an experience's message, remove selected photos, and optionally add more. POST only."""
    if request.method != "POST":
        return redirect("staff_profile", id=id)
    staff = get_object_or_404(Staff.objects.select_related('profile'), id=id)
    if not _can_manage_staff_experiences(request.user, staff):
        messages.error(request, 'You do not have permission to edit experiences on this profile.')
        return redirect('staff_profile', id=id)
    experience = get_object_or_404(StaffExperience, id=experience_id, staff=staff)
    message = request.POST.get("experience", "").strip()
    if message:
        experience.message = message
        experience.save()
        for photo_id in request.POST.getlist("photos_to_remove"):
            try:
                photo = StaffExperiencePhoto.objects.get(id=photo_id, experience=experience)
                photo.delete()
            except (ValueError, StaffExperiencePhoto.DoesNotExist):
                pass
        for photo in experience.photos.all():
            raw_caption = request.POST.get("caption_%s" % photo.id, "").strip()
            new_caption = proper_case(raw_caption) if raw_caption else ""
            if photo.caption != new_caption:
                photo.caption = new_caption
                photo.save()
        captions = request.POST.getlist("photos_captions")
        for i, f in enumerate(request.FILES.getlist("photos")):
            if f.content_type and f.content_type.startswith("image/"):
                raw_caption = captions[i].strip() if i < len(captions) else ""
                caption = proper_case(raw_caption) if raw_caption else ""
                StaffExperiencePhoto.objects.create(experience=experience, image=f, caption=caption)
        messages.success(request, "Experience updated.")
    else:
        messages.warning(request, "Message cannot be empty.")
    return redirect("staff_profile", id=id)


def staff_delete_experience(request, id, experience_id):
    """Delete an experience (and its photos) from a staff profile. POST only."""
    if request.method != "POST":
        return redirect("staff_profile", id=id)
    staff = get_object_or_404(Staff.objects.select_related('profile'), id=id)
    if not _can_manage_staff_experiences(request.user, staff):
        messages.error(request, 'You do not have permission to delete experiences on this profile.')
        return redirect('staff_profile', id=id)
    experience = get_object_or_404(StaffExperience, id=experience_id, staff=staff)
    experience.delete()
    messages.success(request, "Experience deleted.")
    return redirect("staff_profile", id=id)
